from collections import defaultdict
import re
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import random
from deta import Deta

deta = Deta("c09hsnq1_tDcQgDiYGUxUNqFJpMtriV33FU1o1uE5")
drive = deta.Drive('asprak')

def clean_df(raw_df):
    df = raw_df.copy()
    df['Nama Sesuai SSO'] = df['Nama Sesuai SSO'].str.strip()
    df['Nama Sesuai SSO'] = df['Nama Sesuai SSO'].str.title()
    df.columns = list(map(lambda x: x.replace('.', ':'), df.columns))
    df.columns = list(map(lambda x: x.replace('SELESA', 'SELASA'), df.columns))
    df = df.drop_duplicates(subset='Nama Sesuai SSO', keep='last')
    return df

def df_to_schedule(raw_df):
    df = clean_df(raw_df)
    list_nama = df['Nama Sesuai SSO']
    dict_jadwal = {}
    temp_matkul = []
    temp_jadwal = defaultdict(lambda: defaultdict(lambda: {}))
    counter = 0
    for nama in list_nama:
        for key, value in df[df['Nama Sesuai SSO'] == nama].squeeze().items():
            try:
                if value == 'BISA':
                    matkul, hari, jam = re.findall('JAGA\s+(\S+)\s+\[(\S+)\s+(\d{2}[:.]\d{2}\s+-\s+\d{2}[:.]\d{2})', key)[0]
                    temp_jadwal['jadwal'][hari][jam] = None
                    temp_jadwal['appearance'] = 0
                    temp_matkul.append(matkul)
                    counter += 1
            except Exception as e:
                print(str(e))
        temp_jadwal['matkul'] = list(set(temp_matkul))
        temp_jadwal['ketersediaan'] = counter
        dict_jadwal[nama.strip()] = dict(temp_jadwal)
        temp_jadwal = defaultdict(lambda: defaultdict(lambda: {}))
        temp_matkul = []
        counter = 0
    return dict_jadwal

def df_to_excel_schedule(df, judul):
    raw_schedule = df_to_schedule(df)
    asprak_excel = []
    matkul_excel = []
    hari_excel = []
    jam_excel = []
    isi_excel = []
    starts = []
    stops = []
    starts_d = []
    stops_d = []
    temp_start = 2
    temp_stop = 1
    temp_start_d = 2
    temp_stop_d = 1
    for asprak, value in raw_schedule.items():
        matkul_excel.extend(value['matkul'])
        for hari, jadwal in value['jadwal'].items():
            for isi in jadwal.keys():
                asprak_excel.append(asprak)
                matkul_excel.append('')
                hari_excel.append(hari)
                jam_excel.append(isi)
                isi_excel.append('')
                temp_stop += 1
                temp_stop_d += 1
            starts_d.append(temp_start_d)
            stops_d.append(temp_stop_d)
            temp_start_d = temp_stop_d + 1
            temp_stop_d = temp_start_d - 1
        matkul_excel = matkul_excel[:-len(value['matkul'])]
        starts.append(temp_start)
        stops.append(temp_stop)
        temp_start = temp_stop + 1
        temp_stop = temp_start -1
    ready_excel = pd.DataFrame({
        'asprak': asprak_excel,
        'matkul': matkul_excel,
        'hari': hari_excel,
        'jam': jam_excel,
        'isi': isi_excel
    })
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(ready_excel, index=False, header=True):
        ws.append(r)
    for i in range(len(starts)):
        ws.merge_cells(start_row=starts[i], start_column=1, end_row=stops[i], end_column=1)
    
    for i in range(len(starts_d)):
        ws.merge_cells(start_row=starts_d[i], start_column=3, end_row=stops_d[i], end_column=3)

    try:
        os.remove(f'{judul}.xlsx')
    except:
        pass
    wb.save(f'{judul}.xlsx')

def df_to_empty(raw_df):
    df = raw_df.copy()
    jadwal = list(filter(lambda x: x.startswith('JADWAL'), df.columns))
    final_res = defaultdict(lambda: defaultdict(lambda: []))
    for x in jadwal:
        matkul, hari, jam = re.findall('JAGA\s+(\S+)\s+\[(\S+)\s+(\d{2}[:.]\d{2}\s+-\s+\d{2}[:.]\d{2})', x)[0]
        jam = jam.replace('.', ':')
        hari = 'SELASA' if hari == 'SELESA' else hari
        data = {
            'ruangan': None,
            'matkul': matkul,
            'kelas': None,
            'dosen': None,
            'asprak': []
        }
        final_res[hari][jam].append(data)
    return final_res

def assign(raw_df):
    df = raw_df.copy()
    raw_empty = df_to_empty(df)
    raw_schedule = df_to_schedule(df)
    for asprak, value in raw_schedule.items():
        matkul_asprak = value['matkul']
        for hari, praktikum in raw_empty.items():
            jam = praktikum.keys()
            for x in jam:
                for single in praktikum[x]:
                    if single['matkul'] in matkul_asprak and len(single['asprak']) < 6 and value['appearance'] < 5:
                        try:
                            if not value['jadwal'][hari][x]:
                                value['jadwal'][hari][x] = single['matkul']
                                single['asprak'].append(asprak)
                                value['appearance'] += 1
                        except:
                            pass
    return raw_empty

def to_api(empty_jadwal):
    final = []
    for hari, value in empty_jadwal.items():
        for jam, isi in value.items():
            ruang = ['0604', '0605', '0617', '0618', '0704', '0705', '0717', '0718']
            random.shuffle(ruang)
            for jadwal in isi:
                jadwal['hari'] = hari
                jadwal['jam'] = jam
                jadwal['ruangan'] = ruang.pop()
                jadwal['dosen'] = 'NA'
                jadwal['kelas'] = 'NA'
                final.append(jadwal)
    return final
